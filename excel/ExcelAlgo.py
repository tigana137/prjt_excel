from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
import datetime
from datetime import date
from excel.models import excelsheets

from x.models import AdminEcoledata

# list = (column letter , colymn name , data key, column width)
columns = [('A', "رمز المدرسة", "    ", 23),  # 1
           ('B', "المعتمديّة", "Del1", 23),  # 2
           # 3 7attitlha comments 5atr boucle for table row
           ('C', "ع/ر", "comments", 6),
           ('D', "التلميذ(ة)", "nom_prenom", 30),  # 4
           ('E', "المعرف الوحيد", "uid", 19),  # 5
           ('F', "اسم الأب", "nom_pere", 25),  # 6
           ('G', "تاريخ الولادة", "date_naissance", 16),  # 7
           ('H', "المستوى", "level", 12),  # 8
           ('I', "المدرسة المرسم بها", "prev_ecole", 30),  # 9
           ('J', "المدرسة المرغوب فيها", "next_ecole", 30),  # 10
           ('K', "المؤيدات", "reason", 20),  # 11
           ('L', "قرار اللجنة", "decision", 15),  # 12
           ('M', "ملاحظات", "comments", 10)  # 13
           ]

tableHead_row = 7
column_height = 20


def getexcelsheetRows(dre_id):
    data = excelsheets.objects.filter(
        dre_id=dre_id, date_downloaded=None).values()
    return data


def headers(sheet):
    font = Font(size=14, bold=True)
    alignment = Alignment(horizontal='center', vertical='center')
    texts = ['وزارة التربية', 'المندوبية الجهوية للتربية',
             'إدارة المرحلة الابتدائية', 'مصلحة شؤون التلاميذ', 'السنة الدراسية 2024/2023']

    for i in range(1, 5):
        cell = sheet["A"+str(i)]
        sheet.merge_cells("A"+str(i)+":C"+str(i))
        cell.value = texts[int(i)]
        cell.font = font
        cell.alignment = alignment


def get_arabic_date():
    arabic_months = ["", 'جانفي', 'فيفري', 'مارس', 'أفريل', 'ماي',
                     'جوان', 'جويلية', 'أوت', 'سبتمبر', 'أكتوبر', 'نوفمبر', 'ديسمبر']

    current_date = datetime.datetime.now()
    current_month = current_date.month
    current_day = current_date.day
    current_year = current_date.year
    arabic_date = f'{current_day} {arabic_months[current_month]} {current_year}'
    return (arabic_date)


def title(sheet):
    cell = sheet['D3']
    sheet.merge_cells("D3:H4")

    font = Font(size=22, bold=True)
    alignment = Alignment(horizontal='center', vertical='center')

    cell.value = 'مطالب النقل ليوم ' + get_arabic_date()
    cell.font = font
    cell.alignment = alignment


def AdjustWidthHeight(sheet, len_data):
    for (columnLetter, _, _, width) in columns:
        sheet.column_dimensions[columnLetter].width = width

    for i in range(len_data):
        sheet.row_dimensions[i].height = column_height


def tableHead(sheet):

    font = Font(size=11, bold=True)
    alignment = Alignment(horizontal='center', vertical='center')
    color = PatternFill(start_color='FFFF00', fill_type='solid')

    for (columnLetter, title, _, _) in columns:
        cell = sheet[columnLetter+str(tableHead_row)]
        cell.value = title
        cell.font = font
        cell.alignment = alignment
        cell.fill = color


def tableRow(sheet, data):

    font = Font(size=12, bold=True)
    alignment = Alignment(horizontal='center', vertical='center')
    initial_row = tableHead_row+1

    for i in range(len(data)):
        eleve = data[i]

        try:
            cell = sheet["A"+str(initial_row+i)]
            cell.font = font
            cell.alignment = alignment
            cell.value = str(eleve["next_ecole_id"]) + \
                " ⤺ " + str(eleve["prev_ecole_id"])

        except AdminEcoledata.DoesNotExist:
            pass
        for (columnLetter, _, key, _) in columns[1:]:
           # print(column)
            cell = sheet[columnLetter+str(initial_row+i)]
            cell.value = eleve[key]
            cell.font = font
            cell.alignment = alignment
        cell = sheet["C"+str(initial_row+i)]
        cell.value = i+1


def insertTable(sheet, data):
    table_start = 'A'+str(tableHead_row)
    table_end = f'M{len(data)+tableHead_row}'

    # Create a table object
    table = Table(displayName="MyTable", ref=f'{table_start}:{table_end}')

    # Add a TableStyleInfo object to apply table styles
    style = TableStyleInfo(
        name="TableStyleMedium9",  # You can choose a different style if desired
        # showFirstColumn=False,
        # showLastColumn=False,
        # showRowStripes=True,
    )
    table.tableStyleInfo = style

    # Add the table to the worksheet
    sheet.add_table(table)


def getFileName():
    arabic_months = ["", 'جانفي', 'فيفري', 'مارس', 'أفريل', 'ماي',
                     'جوان', 'جويلية', 'أوت', 'سبتمبر', 'أكتوبر', 'نوفمبر', 'ديسمبر']

    current_date = datetime.datetime.now()
    current_month = current_date.month
    current_day = current_date.day
    FileName = f'نقل ليوم  {current_day}{arabic_months[current_month]}'
    return FileName


def archiveExcelsheetRows(dre_id):
    data = excelsheets.objects.filter(dre_id=dre_id, date_downloaded=None)
    for sheetRow in data:
        sheetRow.date_downloaded = date.today()
    excelsheets.objects.bulk_update(data,batch_size=100, fields=["date_downloaded"])


def initiate_Excel(dre_id):
    data = getexcelsheetRows(dre_id)

    workbook = Workbook()       # Create a new workbook to reorder sheets
    sheet = workbook.active
    sheet.title = "تجميع النقل"
    sheet.sheet_view.rightToLeft = True     # Reverse the order of sheet names
    headers(sheet)
    title(sheet)
    AdjustWidthHeight(sheet, len(data))
    tableHead(sheet)
    tableRow(sheet, data)
    insertTable(sheet, data)

    archiveExcelsheetRows(dre_id)

    return workbook, getFileName()
