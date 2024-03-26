

from openpyxl import load_workbook

from x.models import AdminEcoledata

column_starting_point = 'B'
row_starting_point = 8

def annualexcel():
    sids = AdminEcoledata.objects.filter(dre_id=84).values_list('sid',flat=True)
    wb = load_workbook("xx.xlsx")
    ws = wb.active

    row = row_starting_point
    ecoles = []
    while ws['C'+str(row)].value :
        sid = ws['C'+str(row)].value
        if int(sid) not in sids :
            continue
        ministre_school_name = ws['D'+str(row)].value
        premier_elvs = ws['S'+str(row)].value
        premier_classes = ws['T'+str(row)].value
        
        deuxieme_elvs = ws['X'+str(row)].value
        deuxieme_classes = ws['Y'+str(row)].value

        troisieme_elvs = ws['AC'+str(row)].value
        troisieme_classes = ws['AD'+str(row)].value

        quaterieme_elvs = ws['AH'+str(row)].value
        quaterieme_classes = ws['AI'+str(row)].value

        cinquieme_elvs = ws['AM'+str(row)].value
        cinquieme_classes = ws['AN'+str(row)].value

        sixieme_elvs = ws['AR'+str(row)].value
        sixieme_classes = ws['AS'+str(row)].value
        
        AdminEcoledata.objects.get(sid=sid).update_levelstat((premier_elvs,premier_classes),(deuxieme_elvs,deuxieme_classes),(troisieme_elvs,troisieme_classes),(quaterieme_elvs,quaterieme_classes),(cinquieme_elvs,cinquieme_classes),(sixieme_elvs,sixieme_classes))        
        ecole = AdminEcoledata(
            sid=sid,
            ministre_school_name=ministre_school_name
        )
        ecoles.append(ecole)
        row+=1

    AdminEcoledata.objects.bulk_update(ecoles,fields=['ministre_school_name'])
    pass



#  premiere_data.nbr_elvs = premier_elvs,
#           deuxieme_data.nbr_elvs = deuxieme_elvs,
#           troisieme_data.nbr_elvs = troisieme_elvs,
#           quatrieme_data.nbr_elvs = quaterieme_elvs,
#           cinquieme_data.nbr_elvs = cinquieme_elvs,
#           sixieme_data.nbr_elvs = sixieme_elvs,