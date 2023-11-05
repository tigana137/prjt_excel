from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles import PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
import datetime







def headers(sheet):
    font = Font(size=14, bold=True)
    alignment = Alignment(horizontal='center', vertical='center')
    texts = ['وزارة التربية', 'المندوبية الجهوية للتربية',
             'إدارة المرحلة الابتدائية', 'مصلحة شؤون التلاميذ', 'السنة الدراسية 2024/2023']

    for i in range(1,5):
        cell = sheet["A"+str(i)]
        sheet.merge_cells("A"+str(i)+":C"+str(i))
        cell.value = texts[int(i)]
        cell.font = font
        cell.alignment = alignment



def get_arabic_date():
    arabic_months=["",'جانفي','فيفري','مارس','أفريل','ماي','جوان','جويلية','أوت','سبتمبر','أكتوبر','نوفمبر','ديسمبر']

    current_date = datetime.datetime.now()
    current_month = current_date.month
    current_day = current_date.day
    current_year = current_date.year
    arabic_date = f'{current_day} {arabic_months[current_month]} {current_year}'
    return(arabic_date)
    


def title(sheet):
    cell = sheet['D3']
    sheet.merge_cells("D3:H3")

    font = Font(size=16, bold=True)
    alignment = Alignment(horizontal='center', vertical='center')

    cell.value = 'مطالب النقل ليوم '+ get_arabic_date()
    cell.font = font
    cell.alignment = alignment


def AdjustWidthHeight(sheet):
    columns_width = {'A': 14, 'B': 3.5, 'C': 17, 'D': 18, 'E': 21,'F':22,'G':22,'H':22,'I':22,'J':17,'K':10 } 
    for column , width in columns_width.items():
        sheet.column_dimensions[column].width = width
    
    height = 20  
    for i in range(20): # ~ baddlha b lenght t3 l array l chte5dhou mill request front end
        sheet.row_dimensions[i].height = height


def tableHead(sheet):
    titles = ["المعتمديّة","ع/ر","التلميذ(ة)","المعرف الوحيد","اسم الأب","تاريخ الولادة","المدرسة المرسم بها","المدرسة المرغوب فيها","المؤيدات","قرار اللجنة","ملاحظات"]
    row = "5"

    font = Font(size=11, bold=True)
    alignment = Alignment(horizontal='center', vertical='center')
    color = PatternFill(start_color='FFFF00', fill_type='solid')

    for i in range(11):
        column = chr(i+65)
        cell = sheet[column+row]
        cell.value = titles[i]
        cell.font = font
        cell.alignment = alignment
        cell.fill = color
    pass


def tableRow(sheet,data):

    font = Font(size=12, bold=True)
    alignment = Alignment(horizontal='center', vertical='center')
    initial_row = 6
    keys=["Del1","nbr","nom_prenom","uid","nom_pere","date_naissance","prev_ecole","next_ecole","reason","decision" ,"comments"]

    for i in range(len(data)):
        eleve = data[i]
        for k in range(10):
            column = chr(k+65)
           # print(column)
            key = keys[k]
            cell = sheet[column+str(initial_row+i)]
            cell.value = eleve[key] if k!=1 else i+1
            cell.font = font
            cell.alignment = alignment


def insertTable(sheet,data):
    table_start = 'A5'
    table_end = f'K{len(data)+5}'

    # Create a table object
    table = Table(displayName="MyTable", ref=f'{table_start}:{table_end}')

    # Add a TableStyleInfo object to apply table styles
    style = TableStyleInfo(
        name="TableStyleMedium9",  # You can choose a different style if desired
    # showFirstColumn=False,
    # showLastColumn=False,
    # showRowStripes=True,
    )
    table.tableStyleInfo = style

    # Add the table to the worksheet
    sheet.add_table(table)


def getFileName():
    arabic_months=["",'جانفي','فيفري','مارس','أفريل','ماي','جوان','جويلية','أوت','سبتمبر','أكتوبر','نوفمبر','ديسمبر']

    current_date = datetime.datetime.now()
    current_month = current_date.month
    current_day = current_date.day
    FileName = f'نقل ليوم  {current_day}{arabic_months[current_month]}'
    return FileName


def initiate_Excel(data):
    workbook = Workbook()       # Create a new workbook to reorder sheets
    sheet = workbook.active
    sheet.title = "تجميع النقل"
    sheet.sheet_view.rightToLeft = True     # Reverse the order of sheet names
    headers(sheet)
    title(sheet)
    AdjustWidthHeight(sheet)
    tableHead(sheet)
    tableRow(sheet,data)
    insertTable(sheet,data)

    return workbook,getFileName()



